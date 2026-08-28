import os
from django.shortcuts import render, redirect, HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist
from .models import *
from account.models import *
from .forms import MetricsForm, BillForm, ClientForm, BulkUploadForm, CustomerForm
from django.db.models import F, Sum, Q, Count
import sweetify
from account.forms import *
from main.decorators import *
import datetime
from twilio.rest import Client as TwilClient
import csv
import openpyxl
from django.http import HttpResponse, JsonResponse
import json
from django.conf import settings
import stripe
import weasyprint
from account.forms import RegistrationForm
from django.template.loader import render_to_string

stripe.api_key = settings.STRIPE_SECRET_KEY if settings.STRIPE_SECRET_KEY else None

def landingpage(request):
    return render(request, 'landingpage/landingpage.html')  


@staff_required
def export_clients_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="clients.csv"'

    writer = csv.writer(response)
    writer.writerow(['Meter Number', 'First Name', 'Middle Name', 'Last Name', 'Contact Number', 'Address', 'Connection Status'])

    clients = Client.objects.all().values_list('meter_number', 'first_name', 'middle_name', 'last_name', 'contact_number', 'address', 'status')
    for client in clients:
        writer.writerow(client)

    return response


@staff_required
def export_meter_readings_csv(request):
    """Export all meter readings with quote dates to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="meter_readings.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Meter Number', 'Customer Name', 'Quote Date', 'Month', 
        'Previous Reading', 'Current Reading', 'Consumption (cu.m)', 
        'Bill Amount (KES)', 'Penalty (KES)', 'Total Due (KES)', 
        'Payment Status', 'Due Date', 'Connection Status'
    ])

    # Get all bills with related client data
    bills = WaterBill.objects.select_related('name').order_by('-billing_date')
    
    for bill in bills:
        client = bill.name
        writer.writerow([
            client.meter_number if client else 'N/A',
            f"{client.last_name}, {client.first_name}" if client else 'N/A',
            bill.billing_date.strftime('%d-%m-%Y') if bill.billing_date else 'N/A',
            bill.billing_date.strftime('%B %Y') if bill.billing_date else 'N/A',
            bill.previous_reading or '',
            bill.present_reading or '',
            bill.meter_consumption or '',
            bill.compute_bill() or '',
            bill.penalty() or '',
            bill.payable() or '',
            bill.payment_status or 'N/A',
            bill.duedate.strftime('%d-%m-%Y') if bill.duedate else 'N/A',
            client.status if client else 'N/A',
        ])

    return response

@staff_required
def export_ongoing_bills_excel(request):
    """Export ongoing bills to Excel format"""
    from io import BytesIO
    
    # Get ongoing bills
    bills = WaterBill.objects.filter(payment_status='Pending').select_related('name').order_by('-billing_date')
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ongoing Bills"
    
    # Add headers
    headers = ['Name', 'Bill Amount (KES)', 'Meter Consumption (cu.m)', 'Due Date', 
               'Penalty Date', 'Penalty (KES)', 'Payable Amount (KES)', 'Payment Status']
    ws.append(headers)
    
    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    for bill in bills:
        ws.append([
            str(bill.name),
            bill.compute_bill() or 0,
            bill.meter_consumption or 0,
            bill.duedate.strftime('%d-%m-%Y') if bill.duedate else 'N/A',
            bill.penaltydate.strftime('%d-%m-%Y') if bill.penaltydate else 'N/A',
            bill.penalty() or 0,
            bill.payable() or 0,
            bill.payment_status or 'N/A'
        ])
    
    # Adjust column widths
    column_widths = [25, 18, 22, 15, 15, 15, 18, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Center align numeric columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as download
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ongoing_bills.xlsx"'
    
    return response

@staff_required
def export_recent_users_excel(request):
    """Export recent users to Excel format"""
    from io import BytesIO
    
    # Get recent users (same as dashboard)
    recent_users = Account.objects.all().order_by('-created_at')[:10]
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recent Users"
    
    # Add headers
    headers = ['Name', 'Email', 'Status', 'Verified', 'Joined', 'Account Type', 'Active']
    ws.append(headers)
    
    # Style header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    for user in recent_users:
        # Determine approval status
        if user.admin_approved:
            status = 'Approved'
        elif user.rejected:
            status = 'Rejected'
        else:
            status = 'Pending'
        
        # Determine verification status
        verified = 'Yes' if user.verified else 'No'
        
        # Determine account type
        account_type = 'Admin' if (user.is_superuser or user.is_staff) else 'Customer'
        
        # Determine if active
        is_active = 'Yes' if user.is_active else 'No'
        
        ws.append([
            user.get_full_name(),
            user.email,
            status,
            verified,
            user.created_at.strftime('%d-%m-%Y') if user.created_at else 'N/A',
            account_type,
            is_active
        ])
    
    # Adjust column widths
    column_widths = [25, 30, 15, 12, 15, 15, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Center align certain columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as download
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="recent_users.xlsx"'
    
    return response

from django.shortcuts import get_object_or_404
import logging
from django.utils import timezone

logger = logging.getLogger(__name__)

@login_required(login_url='login')
@verified_or_superuser
def download_invoice(request, pk):
    try:
        logger.info(f"Attempting to download invoice for bill ID: {pk}")
        
        # Get the bill or return 404 if not found
        try:
            bill = WaterBill.objects.get(id=pk)
            logger.info(f"Found bill: {bill.id} for {bill.name}")
        except WaterBill.DoesNotExist:
            logger.error(f"Bill with ID {pk} not found in database")
            return render(request, 'main/404.html', {'message': 'The requested bill was not found.'}, status=404)
        except Exception as e:
            logger.error(f"Error retrieving bill {pk}: {str(e)}")
            return render(request, 'main/error.html', 
                        {'message': 'An error occurred while retrieving the bill.', 'error': str(e)}, 
                        status=500)

        # Get rate from Metric model - ensure we get a valid rate
        metric = Metric.objects.first()
        if metric and metric.consump_amount and metric.consump_amount > 0:
            rate = metric.consump_amount
        else:
            rate = 200.0  # Default rate if no valid metric found
        
        # Calculate period (from billing_date to due_date or next month)
        if bill.billing_date:
            period_start = bill.billing_date.strftime('%d %b')
            # Period end is next day or due date - 1 day
            if bill.duedate:
                from dateutil.relativedelta import relativedelta
                period_end_date = bill.duedate - datetime.timedelta(days=1)
            else:
                period_end_date = bill.billing_date + datetime.timedelta(days=27)  # ~monthly
            period_end = period_end_date.strftime('%d %b %Y')
            period = f"{period_start} - {period_end}"
        else:
            period = "N/A"
        
        # Due date
        due_date = bill.duedate.strftime('%d %b %Y') if bill.duedate else "N/A"
        
        # Consumption
        consumption = bill.meter_consumption if bill.meter_consumption is not None else 0
        
        # Amount due (remaining after partial payments)
        amount_due = bill.balance_remaining() if hasattr(bill, 'balance_remaining') else (bill.payable() if hasattr(bill, 'payable') else (consumption * rate))
        
        # Next reading date - exactly one month after billing date
        if bill.billing_date:
            from dateutil.relativedelta import relativedelta
            next_reading = bill.billing_date + relativedelta(months=1)
            next_reading_date = next_reading.strftime('%d %b %Y')
        else:
            next_reading_date = "N/A"

        context = {
            "company_name":     "Timaji Water Services",
            "tagline":          "Pure Water, Pure Life",
            "customer_name":    f"{bill.name.first_name} {bill.name.last_name}",
            "period":           period,
            "due_date":         due_date,
            "previous_reading": f"{bill.previous_reading or 0:04d}",
            "current_reading":  f"{bill.present_reading or 0:04d}",
            "consumption":      consumption,
            "rate":             f"{rate:,.0f}",
            "amount_due":       f"{amount_due:,.2f}",
            "mpesa_paybill":    getattr(settings, "MPESA_SHORTCODE", "") or "",
            "account_number":   bill.name.account_number if bill.name_id else "",
            "mpesa_number":     getattr(settings, "MPESA_SHORTCODE", "") or "",
            "next_reading_date": next_reading_date,
            "email":            "info@timajiwater.co.ke",
            "phone":            "+254 721 974819",
            "auto_print":       request.GET.get('print') == '1' or request.path.endswith('/receipt/'),
        }
        
        # If print/view requested, render HTML receipt directly for clean printing without triggering attachment download
        if request.GET.get('print') == '1' or request.GET.get('view') == '1' or request.path.endswith('/receipt/'):
            return render(request, "main/receipt_template.html", context)

        html_string = render_to_string("main/receipt_template.html", context)
        try:
            pdf_bytes = weasyprint.HTML(string=html_string).write_pdf()
        except TypeError as e:
            # Handle version compatibility issues
            if "PDF.__init__()" in str(e):
                from weasyprint import HTML, CSS
                doc = HTML(string=html_string)
                pdf_bytes = doc.write_pdf()
            else:
                raise
        
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="invoice_{bill.id}.pdf"'
        return response

    except Exception as e:
        logger.error(f"Unexpected error generating PDF for bill {pk}: {str(e)}", exc_info=True)
        return render(request, 'main/error.html', 
                    {'message': 'An unexpected error occurred while generating the PDF.', 'error': str(e)}, 
                    status=500)


@staff_required
def dashboard(request):
    from django.db.models import Count, Q
    
    # Optimize queries with aggregation instead of multiple count() calls
    bills_stats = WaterBill.objects.aggregate(
        total_bills=Count('id'),
        pending_bills=Count('id', filter=Q(payment_status__in=['Pending', 'Partial']))
    )
    
    clients_stats = Client.objects.aggregate(
        connected_clients=Count('id', filter=Q(status='Connected')),
        disconnected_clients=Count('id', filter=Q(status='Disconnected'))
    )
    
    context = {
        'title': 'Dashboard',
        'total_users': Account.objects.filter(is_superuser=False).count(),
        'total_clients': Client.objects.count(),  # Added actual customer count
        'total_bills': bills_stats['total_bills'],
        'pending_bills': bills_stats['pending_bills'],
        'connected_clients': clients_stats['connected_clients'],
        'disconnected_clients': clients_stats['disconnected_clients'],
        'recent_users': Account.objects.filter(is_superuser=False).order_by('-created_at')[:10],
        'ongoingbills': WaterBill.objects.filter(payment_status__in=['Pending', 'Partial']).select_related('name').order_by('-billing_date')[:5],  # Limit to 5 recent bills
    }
    
    return render(request, 'main/dashboard.html', context)

@login_required(login_url='login')
@verified_or_superuser
def ongoing_bills(request):
    if request.user.is_superuser or request.user.is_staff:
        ongoingbills = WaterBill.objects.filter(payment_status__in=['Pending', 'Partial']).select_related('name')
    else:
        ongoingbills = WaterBill.objects.filter(payment_status__in=['Pending', 'Partial'], approval_status='Approved', name__user=request.user).select_related('name')
    context = {
        'title': 'Ongoing Bills',
        'ongoingbills': ongoingbills,
        'form': BillForm()
    }
    if request.method == 'POST':
        billform = BillForm(request.POST)
        if billform.is_valid():
            bill = billform.save()
            sweetify.toast(request, 'Successfully Added.')
            try: 
                receiver = bill.name.contact_number
                print(f"Attempting to send SMS to: {receiver}") # Added for debugging
                totalbill = bill.payable()
                duedate = bill.duedate
                penaltydate = bill.penaltydate
                SID = os.environ.get('TWILIO_ACCOUNT_SID')
                Auth_Token = os.environ.get('TWILIO_AUTH_TOKEN')
                if SID and Auth_Token:
                    sender = '+17262005435'
                    message = f'\n Your Total Bill is: {totalbill} KSH \n\n Your due date is: {duedate} \n\n Your penalty date is: {penaltydate}'
                    cl = TwilClient(SID, Auth_Token)
                    cl.messages.create(body=message, from_=sender, to=receiver)
                    sweetify.toast(request, 'Notification Sent')
                else:
                    sweetify.toast(request, 'Twilio credentials not configured.', icon='warning')
            except Exception as e: # Catch the exception to get more details
                sweetify.toast(request, f'Contact Number is invalid format: {bill.name.contact_number} (Error: {e})', icon='error')
            return HttpResponseRedirect(request.path_info)
        else:
            print(billform.errors) # Add this line to print form errors
            sweetify.toast(request, 'Invalid Details', icon='error')
    return render(request, 'main/billsongoing.html', context)


ONGOINGBILLS_ORDERABLE_DB_FIELDS = {
    'name': 'name__last_name',
    'billing_date': 'billing_date',
    'previous_reading': 'previous_reading',
    'present_reading': 'present_reading',
    'meter_consumption': 'meter_consumption',
    'duedate': 'duedate',
    'penaltydate': 'penaltydate',
    'payment_status': 'payment_status',
}
ONGOINGBILLS_COLUMNS = [
    'name', 'billing_date', 'previous_reading', 'present_reading',
    'meter_consumption', 'compute_bill', 'duedate', 'penaltydate',
    'penalty', 'payable', 'payment_status', 'action',
]


@login_required(login_url='login')
def ongoing_bills_data(request):
    """DataTables server-side processing endpoint for Ongoing Bills."""
    if request.user.is_superuser or request.user.is_staff:
        qs = WaterBill.objects.filter(payment_status__in=['Pending', 'Partial']).select_related('name')
    else:
        qs = WaterBill.objects.filter(
            payment_status__in=['Pending', 'Partial'],
            approval_status='Approved',
            name__user=request.user,
        ).select_related('name')

    records_total = qs.count()

    search_value = request.GET.get('search[value]', '').strip()
    if search_value:
        qs = qs.filter(
            Q(name__first_name__icontains=search_value) |
            Q(name__last_name__icontains=search_value) |
            Q(payment_status__icontains=search_value)
        )
    records_filtered = qs.count()

    order_col_index = request.GET.get('order[0][column]')
    order_dir = request.GET.get('order[0][dir]', 'asc')
    if order_col_index is not None:
        try:
            col_name = ONGOINGBILLS_COLUMNS[int(order_col_index)]
        except (ValueError, IndexError):
            col_name = None
        db_field = ONGOINGBILLS_ORDERABLE_DB_FIELDS.get(col_name)
        if db_field:
            if order_dir == 'desc':
                db_field = f'-{db_field}'
            qs = qs.order_by(db_field)

    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    page = qs[start:start + length] if length != -1 else qs[start:]

    is_staff = request.user.is_superuser or request.user.is_staff
    data = []
    for bill in page:
        edit_link = (
            f'<a href="/bill/update/{bill.id}" class="btn btn-primary btn-sm" title="Edit">'
            f'<i class="fa-regular fa-pen-to-square"></i></a>'
        ) if is_staff else ''
        data.append({
            'name': str(bill.name),
            'billing_date': bill.billing_date.strftime('%B %Y') if bill.billing_date else '',
            'previous_reading': bill.previous_reading,
            'present_reading': bill.present_reading,
            'meter_consumption': f'{bill.meter_consumption} cu.m' if bill.meter_consumption is not None else '',
            'compute_bill': f'KSh {bill.compute_bill()}',
            'duedate': bill.duedate.isoformat() if bill.duedate else '',
            'penaltydate': bill.penaltydate.isoformat() if bill.penaltydate else '',
            'penalty': f'KSh {bill.penalty()}' if bill.penalty() else 'No Penalty',
            'payable': f'KSh {bill.payable()}',
            'payment_status': bill.payment_status,
            'action': (
                f'<button type="button" onclick="printBillReceipt({bill.id})" '
                f'class="btn btn-success btn-sm" title="Print Receipt">'
                f'<i class="fa-solid fa-print"></i></button> ' + edit_link
            ),
        })

    return JsonResponse({
        'draw': int(request.GET.get('draw', 1)),
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data,
    })


@login_required(login_url='login')
@verified_or_superuser
def history_bills(request):
    if request.user.is_superuser or request.user.is_staff:
        billshistory = WaterBill.objects.filter(payment_status__in=['Paid', 'Pending']).select_related('name')
    else:
        billshistory = WaterBill.objects.filter(payment_status__in=['Paid', 'Pending'], approval_status='Approved', name__user=request.user).select_related('name')
    context = {
        'title': 'Bills History',
        'billshistory': billshistory,
    }
    return render(request, 'main/billshistory.html', context)


# Column order MUST match the <thead> in billshistory.html exactly — this is
# the fixed contract between the server-side JSON and the DataTables init.
# Both staff-only columns (Approval Status, second Action) are always
# included here; DataTables hides them for non-staff users client-side via
# `columns.visible`, so the payload shape never changes based on role.
BILLSHISTORY_COLUMNS = [
    'name', 'billing_date', 'previous_reading', 'present_reading',
    'meter_consumption', 'compute_bill', 'duedate', 'penaltydate',
    'penalty', 'payable', 'payment_status', 'approval_status',
    'action', 'staff_action',
]

# Only these are real DB fields DataTables can sort/filter on directly.
# Computed properties (compute_bill, penalty, payable) can't be ordered
# at the DB level without extra work, so they're excluded from sorting.
BILLSHISTORY_ORDERABLE_DB_FIELDS = {
    'name': 'name__last_name',
    'billing_date': 'billing_date',
    'previous_reading': 'previous_reading',
    'present_reading': 'present_reading',
    'meter_consumption': 'meter_consumption',
    'duedate': 'duedate',
    'penaltydate': 'penaltydate',
    'payment_status': 'payment_status',
    'approval_status': 'approval_status',
}


@login_required(login_url='login')
def history_bills_data(request):
    """DataTables server-side processing endpoint for Bills History."""
    if request.user.is_superuser or request.user.is_staff:
        qs = WaterBill.objects.filter(payment_status__in=['Paid', 'Pending']).select_related('name')
    else:
        qs = WaterBill.objects.filter(
            payment_status__in=['Paid', 'Pending'],
            approval_status='Approved',
            name__user=request.user,
        ).select_related('name')

    records_total = qs.count()

    # --- search ---
    search_value = request.GET.get('search[value]', '').strip()
    if search_value:
        qs = qs.filter(
            Q(name__first_name__icontains=search_value) |
            Q(name__last_name__icontains=search_value) |
            Q(payment_status__icontains=search_value) |
            Q(approval_status__icontains=search_value)
        )
    records_filtered = qs.count()

    # --- ordering ---
    order_col_index = request.GET.get('order[0][column]')
    order_dir = request.GET.get('order[0][dir]', 'asc')
    if order_col_index is not None:
        try:
            col_name = BILLSHISTORY_COLUMNS[int(order_col_index)]
        except (ValueError, IndexError):
            col_name = None
        db_field = BILLSHISTORY_ORDERABLE_DB_FIELDS.get(col_name)
        if db_field:
            if order_dir == 'desc':
                db_field = f'-{db_field}'
            qs = qs.order_by(db_field)

    # --- paging ---
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    page = qs[start:start + length] if length != -1 else qs[start:]

    data = []
    for bill in page:
        data.append({
            'name': str(bill.name),
            'billing_date': bill.billing_date.strftime('%B %Y') if bill.billing_date else '',
            'previous_reading': bill.previous_reading,
            'present_reading': bill.present_reading,
            'meter_consumption': f'{bill.meter_consumption} cu.m' if bill.meter_consumption is not None else '',
            'compute_bill': f'KSh {bill.compute_bill()}',
            'duedate': bill.duedate.isoformat() if bill.duedate else '',
            'penaltydate': bill.penaltydate.isoformat() if bill.penaltydate else '',
            'penalty': f'KSh {bill.penalty()}' if bill.penalty() else 'No Penalty',
            'payable': f'KSh {bill.payable()}',
            'payment_status': bill.payment_status,
            'approval_status': bill.approval_status,
            'action': (
                (
                    f'<button type="button" class="btn btn-success btn-sm pay-mpesa-btn shadow-sm" '
                    f'data-bill-id="{bill.id}" data-amount="{bill.payable()}" '
                    f'data-account="{bill.name.account_number if bill.name else ""}" '
                    f'data-phone="{bill.name.contact_number if bill.name else ""}" '
                    f'data-month="{bill.billing_date.strftime("%B %Y") if bill.billing_date else ""}" '
                    f'title="Pay with M-Pesa Express"><i class="fa-solid fa-mobile-screen-button mr-1"></i> Pay with M-Pesa</button> '
                    if bill.payment_status != 'Paid' else ''
                ) +
                f'<a href="/bills/{bill.id}/receipt/" download="invoice_{bill.id}.pdf" '
                f'class="btn btn-secondary btn-sm" title="Download Receipt">'
                f'<i class="fa-solid fa-download"></i></a> '
                f'<button type="button" onclick="printBillReceipt({bill.id})" '
                f'class="btn btn-info btn-sm" title="Print Receipt">'
                f'<i class="fa-solid fa-print"></i></button>'
            ),
            'staff_action': (
                f'<a href="/bill/update/{bill.id}" class="btn btn-primary btn-sm" title="Edit">'
                f'<i class="fa-regular fa-pen-to-square"></i></a> '
                f'<a href="/bill/delete/{bill.id}" class="btn btn-danger btn-sm" title="Delete">'
                f'<i class="fa-solid fa-trash-can"></i></a>'
            ) if (request.user.is_superuser or request.user.is_staff) else '',
        })

    return JsonResponse({
        'draw': int(request.GET.get('draw', 1)),
        'recordsTotal': records_total,
        'recordsFiltered': records_filtered,
        'data': data,
    })

@staff_required
def update_bills(request, pk):
    bill = WaterBill.objects.get(id=pk)
    form = BillForm(instance=bill)
    context = {
        'title': 'Update Bill',
        'bill': bill,
        'form': form,
    }
    if request.method == 'POST':
        form = BillForm(request.POST, instance=bill)
        if form.is_valid():
            print('approval_status in cleaned_data:', form.cleaned_data.get('approval_status'))
            print('approval_status on instance before save:', bill.approval_status)
            bill = form.save()
            print('approval_status on instance after save:', bill.approval_status)
            sweetify.toast(request, f'{bill} updated successfully.')
            return HttpResponseRedirect(reverse('ongoingbills'))
        else:
            print('BillForm errors:', form.errors)
    return render(request, 'main/billupdate.html', context)


@user_passes_test(lambda u: u.is_superuser)
def delete_bills(request, pk):
    bill = WaterBill.objects.get(id=pk)
    context = {
        'title': 'Delete Bill',
        'bill': bill,
    }
    if request.method == 'POST':
        bill.delete()
        sweetify.toast(request, f'{bill} deleted successfully.')
        return HttpResponseRedirect(reverse('ongoingbills'))
    return render(request, 'main/billdelete.html', context)



@login_required(login_url='login')
@verified_or_superuser
def profile(request, pk):
    profile = Account.objects.get(id=pk)
    student_form = UpdateProfileForm(instance=profile)
    if request.method == 'POST':
        student_form = UpdateProfileForm(request.POST, instance=profile)
        password1 = request.POST['password']
        password2 = request.POST['password2']
        if password1 != password2:
            print("password does not match")
            sweetify.error(request, 'Password does not match!')
            return HttpResponseRedirect(request.path_info)
        elif student_form.is_valid():
            student_form.save()
            sweetify.success(request, 'Updated Successfully')
            return HttpResponseRedirect(reverse('login'))
        else: 
            sweetify.error(request, 'Invalid Credentials!')
            return HttpResponseRedirect(request.path_info)
    context = {
        'title': 'Profile',
        'student_form': student_form,
        'profile': profile,
    }
    return render(request, 'main/profile.html', context)

@staff_required
def users_all(request):
    users_list = Account.objects.filter(is_superuser=False)
    search_query = request.GET.get('search', '')
    
    if search_query:
        users_list = users_list.filter(
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(id__icontains=search_query)
        )
    
    context = {
        'title': 'All Users',
        'users': users_list,
        'search_query': search_query
    }
    return render(request, 'main/users.html', context)


@staff_required
def users_pending(request):
    users_list = Account.objects.filter(is_superuser=False, admin_approved=False)
    search_query = request.GET.get('search', '')
    
    if search_query:
        users_list = users_list.filter(
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(id__icontains=search_query)
        )
    
    context = {
        'title': 'Pending Approval',
        'users': users_list,
        'search_query': search_query
    }
    return render(request, 'main/users.html', context)


@staff_required
def users_rejected(request):
    users_list = Account.objects.filter(is_superuser=False, admin_approved=False, is_active=False)
    search_query = request.GET.get('search', '')
    
    if search_query:
        users_list = users_list.filter(
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(id__icontains=search_query)
        )
    
    context = {
        'title': 'Rejected Users',
        'users': users_list,
        'search_query': search_query
    }
    return render(request, 'main/users.html', context)


@staff_required
def users_approved(request):
    users_list = Account.objects.filter(is_superuser=False, admin_approved=True)
    search_query = request.GET.get('search', '')
    
    if search_query:
        users_list = users_list.filter(
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(id__icontains=search_query)
        )
    
    context = {
        'title': 'Approved Users',
        'users': users_list,
        'search_query': search_query
    }
    return render(request, 'main/users.html', context)


@staff_required
def users_active(request):
    users_list = Account.objects.filter(is_superuser=False, is_active=True)
    search_query = request.GET.get('search', '')
    
    if search_query:
        users_list = users_list.filter(
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(id__icontains=search_query)
        )
    
    context = {
        'title': 'Active Users',
        'users': users_list,
        'search_query': search_query
    }
    return render(request, 'main/users.html', context)


@staff_required
def users_inactive(request):
    users_list = Account.objects.filter(is_superuser=False, is_active=False)
    search_query = request.GET.get('search', '')
    
    if search_query:
        users_list = users_list.filter(
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(id__icontains=search_query)
        )
    
    context = {
        'title': 'Inactive Users',
        'users': users_list,
        'search_query': search_query
    }
    return render(request, 'main/users.html', context)


# Keep the old 'users' view for backward compatibility, redirect to users_all
@staff_required
def users(request):
    return redirect('users_all')

@staff_required
def update_user(request, pk):
    user = Account.objects.get(id=pk)
    form = UpdateUserForm(instance=user)
    context = {
        'title': 'Users',
        'user': user,
        'form': form,
    }
    if request.method == 'POST':
        form = UpdateUserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            sweetify.toast(request, f'{user} updated sucessfuly')
            return HttpResponseRedirect(reverse('users'))
    return render(request, 'main/userupdate.html', context)

@staff_required
def view_user_profile(request, pk):
    """View user profile in read-only mode (for admins)"""
    user = Account.objects.get(id=pk)
    
    # Get associated client if exists
    try:
        client = Client.objects.get(user=user)
    except Client.DoesNotExist:
        client = None
    
    context = {
        'title': f'Profile - {user.get_full_name()}',
        'user': user,
        'client': client,
    }
    return render(request, 'main/view_user_profile.html', context)

@staff_required
def print_user_profile(request, pk):
    """Print user profile as a formatted printable document"""
    user = Account.objects.get(id=pk)
    
    # Get associated client if exists
    try:
        client = Client.objects.get(user=user)
    except Client.DoesNotExist:
        client = None
    
    from django.utils import timezone
    context = {
        'title': f'Print Profile - {user.get_full_name()}',
        'user': user,
        'client': client,
        'current_date': timezone.now().strftime('%d %B %Y at %H:%M'),
    }
    return render(request, 'main/user_profile_print.html', context)

@user_passes_test(lambda u: u.is_superuser)
def delete_user(request, pk):
    user = Account.objects.get(id=pk)
    context = {
        'title': 'Users',
        'user': user,
    }
    if request.method == 'POST':
        user.delete()
        sweetify.toast(request, 'Deleted successfuly.')
        return HttpResponseRedirect(reverse('users'))
    return render(request, 'main/userdelete.html', context)

@staff_required
def approve_user(request, pk):
    user = Account.objects.get(id=pk)
    if request.method == 'POST':
        user.admin_approved = True
        user.rejected = False
        user.is_active = True
        user.save()
        sweetify.success(request, f'User {user.email} has been approved successfully.')
        return redirect('users_pending')
    return redirect('users_pending')


@staff_required
def reject_user(request, pk):
    user = Account.objects.get(id=pk)
    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '').strip()
        
        if not rejection_reason:
            sweetify.error(request, 'Please provide a reason for rejection.')
            return redirect('users_pending')
        
        user.admin_approved = False
        user.rejected = True
        user.is_active = False
        user.rejection_reason = rejection_reason
        user.save()
        
        # Create in-app notification
        UserNotification.objects.create(
            user=user,
            notification_type='rejection',
            title='Account Rejection Notice',
            message=f'Your account registration has been rejected. Reason: {rejection_reason}'
        )
        
        # Send email notification
        try:
            import smtplib
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart
            
            SENDER_EMAIL = settings.OTP_EMAIL
            SENDER_PASSWORD = settings.OTP_PASSWORD
            
            if SENDER_EMAIL and SENDER_PASSWORD:
                msg = MIMEMultipart()
                msg['From'] = SENDER_EMAIL
                msg['To'] = user.email
                msg['Subject'] = 'Account Registration Rejection - Water Billing System'
                
                body = f"""
Dear {user.first_name} {user.last_name},

We regret to inform you that your account registration for the Water Billing System has been rejected.

Reason for Rejection:
{rejection_reason}

If you believe this is an error or would like to appeal this decision, please contact our support team.

Thank you for your understanding.

Best regards,
Water Billing System Administration
"""
                msg.attach(MIMEText(body, 'plain'))
                
                server = smtplib.SMTP('smtp.gmail.com', 587)
                server.starttls()
                server.login(SENDER_EMAIL, SENDER_PASSWORD)
                server.sendmail(SENDER_EMAIL, user.email, msg.as_string())
                server.quit()
                
                sweetify.success(request, f'User {user.email} has been rejected and notified via email.')
            else:
                sweetify.success(request, f'User {user.email} has been rejected. (Email notification not configured)')
        except Exception as e:
            print(f"Error sending rejection email: {str(e)}")
            sweetify.success(request, f'User {user.email} has been rejected. (Email notification failed: {str(e)})')
        
        return redirect('users_pending')
    return redirect('users_pending')


@staff_required
def add_user(request):
    form = RegistrationForm()
    context = {
        'title': 'Add User',
        'form': form
    }
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            sweetify.toast(request, 'User added successfully')
            return HttpResponseRedirect(reverse('users'))
        else:
            sweetify.toast(request, 'Invalid details', icon='error')
    return render(request, 'main/useradd.html', context)

@staff_required
def clients(request):
    form = ClientForm()
    context = {
        'title': 'Clients',
        'clients': Client.objects.all(),
        'form': form
    }
    if request.method == 'POST':
        form = ClientForm(request.POST)
        contact_number = request.POST['contact_number']
        check_number = Client.objects.filter(contact_number=contact_number).exists()
        if form.is_valid():
            form.save()
            sweetify.toast(request, 'Client added')
            return HttpResponseRedirect(reverse('clients'))
        elif check_number:
            sweetify.toast(request,'Phone number already exist', icon='error')
        else:
            sweetify.toast(request, 'Invalid details', icon='error')
    return render(request, 'main/clients.html', context)

@staff_required
def client_update(request,pk):
    client = Client.objects.get(id=pk)
    form = ClientForm(instance=client)
    context = {
        'title': 'Update Client',
        'client': client,
        'form': form
    }
    if request.method == 'POST':
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            sweetify.toast(request, 'Client updated successfully')
            return HttpResponseRedirect(reverse('clients'))
        else:
            sweetify.toast(request, 'Invalid Details', icon='error')
    return render(request, 'main/clientupdate.html', context)


@user_passes_test(lambda u: u.is_superuser)
def client_delete(request,pk):
    client = Client.objects.get(id=pk)
    if request.method == 'POST':
        client.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})



@staff_required
def metrics(request):
    clients = Client.objects.all()
    search_query = request.GET.get('search', '')
    
    if search_query:
        clients = clients.filter(
            Q(meter_number__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(id__icontains=search_query)
        )
    
    total_meters = clients.count()
    total_consumption_all = WaterBill.objects.aggregate(Sum('meter_consumption'))['meter_consumption__sum'] or 0
    connected_clients = Client.objects.filter(status='Connected').count()
    disconnected_clients = Client.objects.filter(status='Disconnected').count()
    pending_clients = Client.objects.filter(status='Pending').count()

    for client in clients:
        client.total_consumption = WaterBill.objects.filter(name=client).aggregate(Sum('meter_consumption'))['meter_consumption__sum'] or 0

    context = {
        'title': 'Metrics',
        'clients': clients,
        'total_meters': total_meters,
        'total_consumption': total_consumption_all,
        'connected_clients': connected_clients,
        'disconnected_clients': disconnected_clients,
        'pending_clients': pending_clients,
        'form': CustomerForm(),
        'search_query': search_query
    }
    return render(request, 'main/metrics.html', context)


@staff_required
def metrics_active(request):
    from django.db.models import Sum
    from django.core.paginator import Paginator
    
    clients = Client.objects.filter(status='Connected')
    search_query = request.GET.get('search', '')
    
    if search_query:
        clients = clients.filter(
            Q(meter_number__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(id__icontains=search_query)
        )
    
    total_meters = clients.count()
    total_consumption_all = WaterBill.objects.aggregate(Sum('meter_consumption'))['meter_consumption__sum'] or 0
    
    # Optimize: Use single query with annotation instead of N+1 queries
    clients = clients.annotate(
        total_consumption=Sum('waterbill__meter_consumption')
    ).order_by('id')  # Add ordering for consistent pagination
    
    # Add pagination to prevent loading thousands of records at once
    paginator = Paginator(clients, 50)  # Show 50 clients per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'title': 'Active Meters',
        'clients': page_obj,
        'page_obj': page_obj,
        'total_meters': total_meters,
        'total_consumption': total_consumption_all,
        'connected_clients': total_meters,
        'disconnected_clients': 0,
        'pending_clients': 0,
        'form': CustomerForm(),
        'search_query': search_query
    }
    return render(request, 'main/metrics.html', context)


@staff_required
def metrics_inactive(request):
    from django.db.models import Sum
    from django.core.paginator import Paginator
    
    clients = Client.objects.filter(status__in=['Disconnected', 'Pending'])
    search_query = request.GET.get('search', '')
    
    if search_query:
        clients = clients.filter(
            Q(meter_number__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(id__icontains=search_query)
        )
    
    disconnected_clients = Client.objects.filter(status='Disconnected').count()
    pending_clients = Client.objects.filter(status='Pending').count()
    total_meters = clients.count()
    total_consumption_all = WaterBill.objects.aggregate(Sum('meter_consumption'))['meter_consumption__sum'] or 0
    
    # Optimize: Use single query with annotation instead of N+1 queries
    clients = clients.annotate(
        total_consumption=Sum('waterbill__meter_consumption')
    ).order_by('id')  # Add ordering for consistent pagination
    
    # Add pagination to prevent loading thousands of records at once
    paginator = Paginator(clients, 50)  # Show 50 clients per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'title': 'Inactive Meters',
        'clients': page_obj,
        'page_obj': page_obj,
        'total_meters': total_meters,
        'total_consumption': total_consumption_all,
        'connected_clients': 0,
        'disconnected_clients': disconnected_clients,
        'pending_clients': pending_clients,
        'form': CustomerForm(),
        'search_query': search_query
    }
    return render(request, 'main/metrics.html', context)


@user_passes_test(lambda u: u.is_superuser)
def metrics_add_remove(request):
    from django.db.models import Sum
    from django.core.paginator import Paginator
    
    clients = Client.objects.all()
    search_query = request.GET.get('search', '')
    
    if search_query:
        clients = clients.filter(
            Q(meter_number__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(id__icontains=search_query)
        )
    
    total_meters = clients.count()
    total_consumption_all = WaterBill.objects.aggregate(Sum('meter_consumption'))['meter_consumption__sum'] or 0
    connected_clients = Client.objects.filter(status='Connected').count()
    disconnected_clients = Client.objects.filter(status='Disconnected').count()
    pending_clients = Client.objects.filter(status='Pending').count()

    # Optimize: Use single query with annotation instead of N+1 queries
    clients = clients.annotate(
        total_consumption=Sum('waterbill__meter_consumption')
    ).order_by('id')  # Add ordering for consistent pagination
    
    # Add pagination to prevent loading thousands of records at once
    paginator = Paginator(clients, 50)  # Show 50 clients per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'title': 'Add/Remove Meters',
        'clients': page_obj,
        'page_obj': page_obj,
        'total_meters': total_meters,
        'total_consumption': total_consumption_all,
        'connected_clients': connected_clients,
        'disconnected_clients': disconnected_clients,
        'pending_clients': pending_clients,
        'form': CustomerForm(),
        'search_query': search_query,
        'show_add_remove': True
    }
    return render(request, 'main/metrics_add_remove.html', context)


@user_passes_test(lambda u: u.is_superuser)
def assign_meter(request):
    all_clients = Client.objects.all()
    search_query = request.GET.get('search', '')
    
    if search_query:
        all_clients = all_clients.filter(
            Q(meter_number__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(id__icontains=search_query)
        )
    
    context = {
        'title': 'Assign Meter',
        'clients': all_clients,
        'search_query': search_query
    }
    
    if request.method == 'POST':
        client_id = request.POST.get('client_id')
        meter_number = request.POST.get('meter_number')
        
        if client_id and meter_number:
            try:
                client = Client.objects.get(id=client_id)
                # Check if meter number is already assigned
                if Client.objects.filter(meter_number=meter_number).exclude(id=client_id).exists():
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'message': 'Meter number already assigned to another client.'})
                    sweetify.error(request, 'Meter number already assigned to another client.')
                else:
                    client.meter_number = meter_number
                    client.save()
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': True, 'message': 'Meter assigned successfully.'})
                    sweetify.success(request, 'Meter assigned successfully.')
                    return redirect('assign_meter')
            except Client.DoesNotExist:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': 'Client not found.'})
                sweetify.error(request, 'Client not found.')
    
    return render(request, 'main/assign_meter.html', context)


@user_passes_test(lambda u: u.is_superuser)
def meters_map(request):
    filter_type = request.GET.get('filter', 'all')  # all, active, inactive
    clients = Client.objects.exclude(latitude__isnull=True).exclude(longitude__isnull=True)
    
    if filter_type == 'active':
        clients = clients.filter(status='Connected')
    elif filter_type == 'inactive':
        clients = clients.filter(status__in=['Disconnected', 'Pending'])
    
    # Calculate total consumption for each client
    for client in clients:
        client.total_consumption = WaterBill.objects.filter(name=client).aggregate(Sum('meter_consumption'))['meter_consumption__sum'] or 0
    
    # Convert to JSON for map
    meters_data = []
    for client in clients:
        if client.latitude and client.longitude:
            meters_data.append({
                'id': client.id,
                'meter_number': client.meter_number or 'N/A',
                'name': f"{client.first_name} {client.last_name}",
                'address': client.address,
                'status': client.status,
                'latitude': float(client.latitude),
                'longitude': float(client.longitude),
                'total_consumption': client.total_consumption,
                'contact_number': client.contact_number or 'N/A',
            })
    
    context = {
        'title': 'Meters Map',
        'meters_data': json.dumps(meters_data),
        'filter_type': filter_type,
        'total_meters': len(meters_data),
        'active_count': Client.objects.filter(status='Connected').exclude(latitude__isnull=True).exclude(longitude__isnull=True).count(),
        'inactive_count': Client.objects.filter(status__in=['Disconnected', 'Pending']).exclude(latitude__isnull=True).exclude(longitude__isnull=True).count(),
    }
    return render(request, 'main/meters_map.html', context)


@user_passes_test(lambda u: u.is_superuser)
def metrics_update(request, pk):
    metric = Metric.objects.get(id=pk)
    form = MetricsForm(instance=metric)
    context = {
        'title': 'Update Metrics',
        'metric': metric,
        'form': form,
    }
    if request.method == 'POST':
        form = MetricsForm(request.POST, instance=metric)
        if form.is_valid():
            form.save()
            sweetify.toast(request, 'Metrics updated successfully.')
            return HttpResponseRedirect(reverse('metrics'))
    return render(request, 'main/metricsupdate.html', context)


@staff_required
def add_customer(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            try:
                client = form.save()
                sweetify.success(request, f'Customer {client.first_name} {client.last_name} added successfully!')
                return JsonResponse({'success': True, 'message': 'Customer added successfully!'})
            except Exception as e:
                print(f"Error saving customer: {str(e)}")
                return JsonResponse({'success': False, 'message': f'Error saving customer: {str(e)}', 'form_html': render_to_string('main/customer_form_partial.html', {'form': form}, request=request)})
        else:
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            print(f"Form validation errors: {error_messages}")
            return JsonResponse({
                'success': False, 
                'message': 'Please correct the errors below: ' + '; '.join(error_messages),
                'form_html': render_to_string('main/customer_form_partial.html', {'form': form}, request=request)
            })
    else:
        form = CustomerForm()
    return render(request, 'main/customer_form_partial.html', {'form': form})


@staff_required
def edit_customer(request, pk):
    try:
        client = Client.objects.get(id=pk)
    except Client.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Client not found.'}, status=404)

    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=client)
        if form.is_valid():
            try:
                client = form.save()
                sweetify.success(request, f'Customer {client.first_name} {client.last_name} updated successfully!')
                return JsonResponse({'success': True, 'message': 'Customer updated successfully!'})
            except Exception as e:
                print(f"Error updating customer: {str(e)}")
                return JsonResponse({'success': False, 'message': f'Error updating customer: {str(e)}', 'form_html': render_to_string('main/customer_form_partial.html', {'form': form, 'client': client}, request=request)})
        else:
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            return JsonResponse({
                'success': False, 
                'message': 'Please correct the errors below.',
                'form_html': render_to_string('main/customer_form_partial.html', {'form': form, 'client': client}, request=request)
            })
    else: # GET request
        form = CustomerForm(instance=client)

    context = {
        'form': form,
        'client': client,
    }
    return render(request, 'main/customer_form_partial.html', context)



@staff_required
def bulk_upload_view(request):
    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['csv_file']
            filename = uploaded_file.name.lower()
            
            try:
                readings_data = []
                
                if filename.endswith('.csv'):
                    decoded_file = uploaded_file.read().decode('utf-8').splitlines()
                    reader = csv.DictReader(decoded_file)
                    for row in reader:
                        readings_data.append(row)
                
                elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                    sheet = wb.active
                    headers = [cell.value for cell in sheet[1]]
                    header_map = {header.strip(): i for i, header in enumerate(headers) if header}
                    
                    for row_idx in range(2, sheet.max_row + 1):
                        row_vals = [cell.value for cell in sheet[row_idx]]
                        if not any(row_vals): continue
                        
                        readings_data.append({
                            'meter_number': row_vals[header_map.get('meter_number')] if 'meter_number' in header_map else None,
                            'billing_date': row_vals[header_map.get('billing_date')] if 'billing_date' in header_map else None,
                            'previous_reading': row_vals[header_map.get('previous_reading')] if 'previous_reading' in header_map else None,
                            'present_reading': row_vals[header_map.get('present_reading')] if 'present_reading' in header_map else None,
                            'due_date': row_vals[header_map.get('due_date')] if 'due_date' in header_map else None,
                            'penalty_date': row_vals[header_map.get('penalty_date')] if 'penalty_date' in header_map else None,
                        })
                else:
                    sweetify.error(request, 'Unsupported file format. Please upload CSV or Excel file.')
                    return redirect('bulk_upload')

                count = 0
                for row in readings_data:
                    try:
                        if not row.get('meter_number'): continue
                        
                        client = Client.objects.get(meter_number=row['meter_number'])
                        WaterBill.objects.create(
                            name=client,
                            billing_date=row['billing_date'],
                            previous_reading=row['previous_reading'],
                            present_reading=row['present_reading'],
                            duedate=row.get('due_date'),
                            penaltydate=row.get('penalty_date'),
                            payment_status='Pending'
                        )
                        count += 1
                    except Client.DoesNotExist:
                        logger.warning(f"Client with meter number {row.get('meter_number')} does not exist.")
                    except Exception as e:
                        logger.error(f"Error processing row: {e}")

                sweetify.success(request, f'Successfully uploaded {count} meter readings.')
                return HttpResponseRedirect(reverse('ongoingbills'))
            except Exception as e:
                sweetify.error(request, f"An error occurred: {e}")
                return redirect('bulk_upload')
    else:
        form = BulkUploadForm()
    return render(request, 'main/bulk_upload.html', {'form': form})


@user_passes_test(lambda u: u.is_superuser)
def bulk_upload_users_view(request):
    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['csv_file']
            filename = uploaded_file.name.lower()
            
            try:
                users_data = []
                
                if filename.endswith('.csv'):
                    decoded_file = uploaded_file.read().decode('utf-8').splitlines()
                    reader = csv.DictReader(decoded_file)
                    for row in reader:
                        users_data.append({
                            'first_name': row.get('First Name'),
                            'last_name': row.get('Last Name'),
                            'email': row.get('Email'),
                            'contact_number': row.get('Contact Number'),
                            'address': row.get('Address')
                        })
                
                elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                    sheet = wb.active
                    
                    # Get headers from first row
                    headers = [cell.value for cell in sheet[1]]
                    
                    # Map header names to indices
                    header_map = {}
                    for i, header in enumerate(headers):
                        if header:
                            header_map[header.strip()] = i
                    
                    # Parse data rows
                    for row_idx in range(2, sheet.max_row + 1):
                        row = [cell.value for cell in sheet[row_idx]]
                        if not any(row): continue # Skip empty rows
                        
                        users_data.append({
                            'first_name': row[header_map.get('First Name')] if 'First Name' in header_map else None,
                            'last_name': row[header_map.get('Last Name')] if 'Last Name' in header_map else None,
                            'email': row[header_map.get('Email')] if 'Email' in header_map else None,
                            'contact_number': row[header_map.get('Contact Number')] if 'Contact Number' in header_map else None,
                            'address': row[header_map.get('Address')] if 'Address' in header_map else None
                        })
                
                else:
                    sweetify.error(request, 'Unsupported file format. Please upload CSV or Excel file.')
                    return redirect('bulk_upload_users')

                count = 0
                for data in users_data:
                    email = data.get('email')
                    if not email:
                        continue
                        
                    if Account.objects.filter(email=email).exists():
                        continue
                        
                    # Create Account
                    user = Account.objects.create_user(
                        email=email,
                        password='Welcome123', # Default password
                        first_name=data.get('first_name'),
                        last_name=data.get('last_name'),
                        admin_approved=True,
                        verified=True
                    )
                    
                    # Create Client
                    Client.objects.create(
                        user=user,
                        first_name=data.get('first_name'),
                        last_name=data.get('last_name'),
                        contact_number=data.get('contact_number'),
                        address=data.get('address'),
                        status='Connected'
                    )
                    count += 1
                
                sweetify.success(request, f'Successfully uploaded {count} users.')
                return redirect('users_all')
            except Exception as e:
                sweetify.error(request, f'An error occurred: {str(e)}')
                return redirect('bulk_upload_users')
    else:
        form = BulkUploadForm()
    return render(request, 'main/bulk_upload_users.html', {'form': form, 'title': 'Bulk Upload Users'})


@staff_required
def send_reminders_view(request):
    if request.method == 'POST':
        ongoing_bills = WaterBill.objects.filter(payment_status='Pending')
        for bill in ongoing_bills:
            try:
                SID = os.environ.get('TWILIO_ACCOUNT_SID')
                Auth_Token = os.environ.get('TWILIO_AUTH_TOKEN')
                if SID and Auth_Token:
                    sender = '+17262005435'
                    receiver = bill.name.contact_number
                    message = f'\n Your Total Bill is: {bill.total_bill} KSH \n\n Your due date is: {bill.due_date} \n\n Your penalty date is: {bill.penalty_date}'
                    cl = TwilClient(SID, Auth_Token)
                    cl.messages.create(body=message, from_=sender, to=receiver)
                    sweetify.toast(request, f'Reminder sent to {bill.name.first_name} {bill.name.last_name}')
                else:
                    sweetify.toast(request, 'Twilio credentials not configured.', icon='warning')
                    break # Stop sending reminders if keys are not set
            except:
                sweetify.toast(request, f'Could not send reminder to {bill.name.first_name} {bill.name.last_name}', icon='error')
        return HttpResponseRedirect(reverse('ongoingbills'))
    return render(request, 'main/send_reminders.html')


@staff_required
def approve_bills_view(request):
    bills = WaterBill.objects.filter(approval_status='Pending Approval')
    context = {
        'title': 'Approve Bills',
        'bills': bills
    }
    return render(request, 'main/approve_bills.html', context)

@staff_required
def bill_approve(request, pk):
    bill = WaterBill.objects.get(id=pk)
    bill.approval_status = 'Approved'
    bill.save()
    sweetify.toast(request, 'Bill approved successfully')
    return HttpResponseRedirect(reverse('approve_bills'))

@staff_required
def bill_reject(request, pk):
    bill = WaterBill.objects.get(id=pk)
    bill.approval_status = 'Rejected'
    bill.save()
    sweetify.toast(request, 'Bill rejected successfully')
    return HttpResponseRedirect(reverse('approve_bills'))



@login_required(login_url='login')
def usage_analytics_view(request):
    from django.db.models import Sum, Count
    from django.db.models.functions import TruncMonth
    
    if request.user.is_superuser or request.user.is_staff:
        # Aggregate by month instead of loading all individual bills
        monthly_data = WaterBill.objects.exclude(billing_date__isnull=True).annotate(
            month=TruncMonth('billing_date')
        ).values('month').annotate(
            total_consumption=Sum('meter_consumption'),
            bill_count=Count('id')
        ).order_by('month')
    else:
        # For regular users, show only their data aggregated by month
        monthly_data = WaterBill.objects.filter(
            name__user=request.user
        ).exclude(billing_date__isnull=True).annotate(
            month=TruncMonth('billing_date')
        ).values('month').annotate(
            total_consumption=Sum('meter_consumption'),
            bill_count=Count('id')
        ).order_by('month')

    labels = [item['month'].strftime('%B %Y') for item in monthly_data]
    data = [item['total_consumption'] or 0 for item in monthly_data]

    average_usage = sum(data) / len(data) if data else 0
    highest_consumption = max(data) if data else 0
    highest_consumption_month = labels[data.index(highest_consumption)] if data else 'N/A'

    context = {
        'title': 'Usage Analytics',
        'labels': labels,
        'data': data,
        'average_usage': average_usage,
        'highest_consumption': highest_consumption,
        'highest_consumption_month': highest_consumption_month,
    }
    return render(request, 'main/usage_analytics.html', context)


@login_required(login_url='login')
def create_checkout_session(request, pk):
    if not settings.STRIPE_SECRET_KEY:
        sweetify.toast(request, 'Stripe is not configured.', icon='error')
        return HttpResponseRedirect(reverse('ongoingbills'))
    bill = WaterBill.objects.get(id=pk)
    session = stripe.checkout.Session.create(
        payment_method_types=['card'],
        line_items=[{
            'price_data': {
                'currency': 'kes',
                'product_data': {
                    'name': f'Water Bill for {bill.name.first_name} {bill.name.last_name}',
                },
                'unit_amount': int(bill.payable() * 100),
            },
            'quantity': 1,
        }],
        mode='payment',
        success_url=request.build_absolute_uri(reverse('payment_success', args=[bill.id])),
        cancel_url=request.build_absolute_uri(reverse('payment_cancel')),
    )
    return redirect(session.url, code=303)


@login_required(login_url='login')
def payment_success(request, pk):
    bill = WaterBill.objects.get(id=pk)
    bill.payment_status = 'Paid'
    bill.save()
    sweetify.toast(request, 'Payment successful.')
    return HttpResponseRedirect(reverse('billshistory'))


@login_required(login_url='login')
def payment_cancel(request):
    sweetify.toast(request, 'Payment cancelled.', icon='error')
    return HttpResponseRedirect(reverse('ongoingbills'))


@login_required(login_url='login')
@verified_or_superuser
def user_dashboard(request):
    try:
        client = Client.objects.get(user=request.user)
        bills_qs = WaterBill.objects.filter(name=client, approval_status='Approved', payment_status__in=['Paid', 'Pending', 'Partial']).order_by('-billing_date')
        bills = [
            {
                'billing_date': bill.billing_date.strftime('%Y-%m-%d') if bill.billing_date else '',
                'previous_reading': bill.previous_reading,
                'present_reading': bill.present_reading,
                'meter_consumption': bill.meter_consumption,
                'payment_status': bill.payment_status,
            }
            for bill in bills_qs
        ]
        metrics = Metric.objects.get(user=request.user)
        print(f"Client found: {client.first_name} {client.last_name}, Meter: {client.meter_number}")
        print(f"Metrics found: Consumption Amount: {metrics.consump_amount}, Penalty Amount: {metrics.penalty_amount}")
    except Client.DoesNotExist:
        client = None
        bills = []
        metrics = None
        print("Client does not exist for this user.")
    except Metric.DoesNotExist:
        metrics = None
        print("Metrics do not exist for this user.")

    context = {
        'title': 'My Meter',
        'client': client,
        'bills': bills,  # list of dicts for chart.js
        'bills_qs': bills_qs if client else [],  # queryset for table display
        'metrics': metrics,
    }
    return render(request, 'main/user_dashboard.html', context)


@login_required(login_url='login')
@verified_or_superuser
def notifications_view(request):
    """User notifications page"""
    user = request.user
    notifications = UserNotification.objects.filter(user=user).order_by('-created_at')
    unread_count = notifications.filter(is_read=False).count()
    
    # Mark all as read if requested
    if request.GET.get('mark_all_read') == 'true':
        from django.utils import timezone
        notifications.filter(is_read=False).update(is_read=True, read_at=timezone.now())
        sweetify.success(request, 'All notifications marked as read.')
        return redirect('notifications')
    
    context = {
        'title': 'Notifications',
        'notifications': notifications,
        'unread_count': unread_count,
    }
    return render(request, 'main/notifications.html', context)


@login_required(login_url='login')
@verified_or_superuser
def mark_notification_read(request, pk):
    """Mark a notification as read"""
    try:
        notification = UserNotification.objects.get(id=pk, user=request.user)
        notification.mark_as_read()
        return JsonResponse({'success': True})
    except UserNotification.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Notification not found'}, status=404)


@login_required(login_url='login')
@verified_or_superuser
def settings_view(request):
    """Settings page for both admin and regular users"""
    user = request.user
    profile_form = UpdateProfileForm(instance=user)
    support_tickets = SupportTicket.objects.filter(user=user).order_by('-created_at')[:10]
    
    if request.method == 'POST':
        if 'update_profile' in request.POST:
            profile_form = UpdateProfileForm(request.POST, instance=user)
            password1 = request.POST.get('password', '')
            password2 = request.POST.get('password2', '')
            
            if password1 and password1 != password2:
                sweetify.error(request, 'Passwords do not match!')
                return redirect('settings')
            elif profile_form.is_valid():
                profile_form.save()
                sweetify.success(request, 'Profile updated successfully!')
                return redirect('settings')
            else:
                sweetify.error(request, 'Invalid form data. Please check your inputs.')
                return redirect('settings')
    
    context = {
        'title': 'Settings',
        'profile_form': profile_form,
        'user': user,
        'support_tickets': support_tickets,
    }
    return render(request, 'main/settings.html', context)


@login_required(login_url='login')
@verified_or_superuser
def user_support(request):
    """User view for their support tickets"""
    user = request.user
    tickets = SupportTicket.objects.filter(user=user).order_by('-created_at')
    
    context = {
        'title': 'Support',
        'tickets': tickets,
    }
    return render(request, 'main/user_support.html', context)


@login_required(login_url='login')
@verified_or_superuser
def contact_support(request):
    """Contact support - create a support ticket"""
    if request.method == 'POST':
        subject = request.POST.get('subject', '')
        message = request.POST.get('message', '')
        priority = request.POST.get('priority', 'Medium')
        
        if subject and message:
            ticket = SupportTicket.objects.create(
                user=request.user,
                subject=subject,
                message=message,
                priority=priority
            )
            sweetify.success(request, f'Support ticket #{ticket.id} created successfully. We will get back to you soon.')
            return redirect('user_support')
        else:
            sweetify.error(request, 'Please fill in both subject and message.')
            return redirect('user_support')
    
    return redirect('user_support')


@staff_required
def support_tickets(request):
    """Admin view for managing support tickets"""
    status_filter = request.GET.get('status', '')
    priority_filter = request.GET.get('priority', '')
    
    tickets = SupportTicket.objects.all()
    
    if status_filter:
        tickets = tickets.filter(status=status_filter)
    if priority_filter:
        tickets = tickets.filter(priority=priority_filter)
    
    tickets = tickets.order_by('-created_at')
    
    context = {
        'title': 'Support Tickets',
        'tickets': tickets,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
    }
    return render(request, 'main/support_tickets.html', context)


@staff_required
def update_ticket(request, pk):
    """Admin view to update support ticket status and add response"""
    ticket = SupportTicket.objects.get(id=pk)
    
    if request.method == 'POST':
        status = request.POST.get('status', ticket.status)
        admin_response = request.POST.get('admin_response', '')
        
        ticket.status = status
        if admin_response:
            ticket.admin_response = admin_response
        if status in ['Resolved', 'Closed'] and not ticket.resolved_at:
            from django.utils import timezone
            ticket.resolved_at = timezone.now()
        ticket.save()
        
        sweetify.success(request, f'Ticket #{ticket.id} updated successfully.')
        return redirect('support_tickets')
    
    return redirect('support_tickets')


@staff_required
def meter_readings_dashboard(request):
    """
    Admin dashboard for viewing and managing meter readings
    Displays all customers with their recent readings and billing info
    """
    from django.shortcuts import get_object_or_404
    from django.core.paginator import Paginator
    from django.db.models import Max, Subquery, OuterRef
    
    # Get all clients
    clients = Client.objects.select_related('user').all()
    
    # Get search/filter parameters
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    if search_query:
        clients = clients.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(meter_number__icontains=search_query) |
            Q(contact_number__icontains=search_query)
        )
    
    if status_filter:
        clients = clients.filter(status=status_filter)
    
    # Optimize: Use subquery to get latest bill for each client in a single query
    latest_bill_subquery = WaterBill.objects.filter(
        name=OuterRef('pk')
    ).order_by('-billing_date')
    
    clients = clients.annotate(
        latest_bill_id=Subquery(latest_bill_subquery.values('id')[:1])
    ).order_by('id')  # Add ordering for consistent pagination
    
    # Add pagination to prevent loading thousands of records at once
    paginator = Paginator(clients, 50)  # Show 50 clients per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Get all latest bills in a single query
    latest_bill_ids = [client.latest_bill_id for client in page_obj if client.latest_bill_id]
    latest_bills = {bill.id: bill for bill in WaterBill.objects.filter(id__in=latest_bill_ids).select_related('name')}
    
    # Build client data with latest readings
    client_data = []
    for client in page_obj:
        latest_bill = latest_bills.get(client.latest_bill_id) if client.latest_bill_id else None
        
        data = {
            'client': client,
            'meter_number': client.meter_number,
            'name': f"{client.last_name}, {client.first_name}",
            'status': client.status,
            'latest_bill': latest_bill,
            'previous_reading': latest_bill.previous_reading if latest_bill else None,
            'current_reading': latest_bill.present_reading if latest_bill else None,
            'consumption': latest_bill.meter_consumption if latest_bill else None,
            'bill_amount': latest_bill.compute_bill() if latest_bill else None,
            'month': latest_bill.billing_date.strftime('%B %Y') if latest_bill else 'N/A',
            'billing_date': latest_bill.billing_date if latest_bill else None,
            'payment_status': latest_bill.payment_status if latest_bill else 'N/A',
        }
        client_data.append(data)
    
    context = {
        'title': 'Meter Readings Dashboard',
        'client_data': client_data,
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'form': BillForm(),
    }
    
    return render(request, 'main/meter_readings_dashboard.html', context)


@staff_required
def add_meter_reading(request, client_id):
    """
    Add a new meter reading for a specific customer
    """
    from django.shortcuts import get_object_or_404
    
    client = get_object_or_404(Client, id=client_id)
    
    if request.method == 'POST':
        form = BillForm(request.POST)
        if form.is_valid():
            bill = form.save()
            sweetify.success(request, f'Meter reading added for {client}')
            return HttpResponseRedirect(reverse('meter_readings_dashboard'))
        else:
            sweetify.error(request, f'Error adding reading')
            return HttpResponseRedirect(reverse('meter_readings_dashboard'))
    else:
        # Pre-fill customer and get last reading
        last_bill = WaterBill.objects.filter(name=client).order_by('-billing_date').first()
        initial_data = {
            'name': client,
            'previous_reading': last_bill.present_reading if last_bill else 0,
            'payment_status': 'Pending',
        }
        form = BillForm(initial=initial_data)
    
    context = {
        'title': f'Add Meter Reading - {client}',
        'client': client,
        'form': form,
    }
    
    return render(request, 'main/add_meter_reading.html', context)


@staff_required
def customer_reading_history(request, client_id):
    """
    View complete meter reading history for a customer
    """
    from django.shortcuts import get_object_or_404
    from django.core.paginator import Paginator
    
    client = get_object_or_404(Client, id=client_id)
    bills = WaterBill.objects.filter(name=client).select_related('name').order_by('-billing_date')
    
    # Calculate statistics using aggregation
    paid_bills_count = bills.filter(payment_status='Paid').count()
    
    # Add pagination to prevent loading hundreds of bills at once
    paginator = Paginator(bills, 50)  # Show 50 bills per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'title': f'Reading History - {client}',
        'client': client,
        'bills': page_obj,
        'page_obj': page_obj,
        'paid_bills_count': paid_bills_count,
    }
    
    return render(request, 'main/customer_reading_history.html', context)


@login_required(login_url='login')
@verified_or_superuser
def customer_bills_view(request):
    """
    Customer-facing view to see their bills (read-only)
    Shows all approved bills with download receipt option
    """
    try:
        client = Client.objects.get(user=request.user)
        # Get all approved bills for this customer (both paid and pending)
        bills = WaterBill.objects.filter(
            name=client, 
            approval_status='Approved'
        ).order_by('-billing_date')
        
        # Build bill data with calculations
        bill_data = []
        for bill in bills:
            data = {
                'id': bill.id,
                'month': bill.billing_date.strftime('%B %Y') if bill.billing_date else 'N/A',
                'billing_date': bill.billing_date,
                'previous_reading': bill.previous_reading,
                'present_reading': bill.present_reading,
                'consumption': bill.meter_consumption,
                'bill_amount': bill.compute_bill(),
                'penalty': bill.penalty(),
                'total': bill.payable(),
                'payment_status': bill.payment_status,
                'due_date': bill.duedate,
                'penalty_date': bill.penaltydate,
            }
            bill_data.append(data)
        
    except Client.DoesNotExist:
        client = None
        bills = []
        bill_data = []
    
    context = {
        'title': 'My Bills',
        'client': client,
        'bills': bills,
        'bill_data': bill_data,
    }
    return render(request, 'main/customer_bills.html', context)
